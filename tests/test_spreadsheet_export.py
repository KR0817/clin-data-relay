from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.spreadsheet_export import ArtifactToolSpreadsheetExporter


def test_openpyxl_fallback_exports_safe_workbook_without_node(tmp_path: Path) -> None:
    exporter = ArtifactToolSpreadsheetExporter(
        node_executable=None,
        builder_script=tmp_path / "missing-builder.mjs",
        output_directory=tmp_path / "exports",
    )

    assert exporter.ready is True

    workbook_bytes = exporter.export(
        {
            "generated_at": "2026-08-10T00:00:00Z",
            "scope": "SITE_A",
            "dictionary_id": "dict-1",
            "dictionary_version": "1.0",
            "submitted_value_count": 1,
            "events": {
                "WEEK_0": {
                    "columns": [{"field_code": "ALT", "display_header": "ALT"}],
                    "rows": [
                        {
                            "centre_code": "SITE_A",
                            "edc_subject_ref": "SYN-001",
                            "values": {"ALT": "=HYPERLINK(\"https://example.test\")"},
                        }
                    ],
                }
            },
            "field_mapping": [
                {
                    "event_ref": "WEEK_0",
                    "field_code": "ALT",
                    "display_header": "ALT",
                    "source_header": "丙氨酸氨基转移酶",
                    "revision": 1,
                }
            ],
        }
    )

    assert workbook_bytes.startswith(b"PK")
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    assert workbook.sheetnames == ["导出说明", "WEEK_0", "字段映射"]
    assert workbook["WEEK_0"]["C2"].value == "'=HYPERLINK(\"https://example.test\")"
    assert workbook["WEEK_0"]["C2"].data_type == "s"
    assert workbook["字段映射"]["D2"].value == "丙氨酸氨基转移酶"


def test_openpyxl_reviewed_export_labels_authority_state_and_only_requested_columns(
    tmp_path: Path,
) -> None:
    exporter = ArtifactToolSpreadsheetExporter(
        node_executable=None,
        builder_script=tmp_path / "missing-builder.mjs",
        output_directory=tmp_path / "exports",
    )

    workbook_bytes = exporter.export(
        {
            "generated_at": "2026-08-11T00:00:00Z",
            "scope": "SITE_A",
            "dictionary_id": "dict-1",
            "dictionary_version": "1.0",
            "export_kind": "reviewed_recognition",
            "export_title": "已确认识别数据导出",
            "inclusion_rule": "仅纳入已人工确认的识别候选",
            "authority_note": "未提交值不是 LibreClinica 权威记录",
            "reviewed_value_count": 1,
            "value_count": 1,
            "include_authority_status": True,
            "authority_status_header": "外部EDC状态",
            "events": {
                "WEEK_0": {
                    "columns": [{"field_code": "PFT_FVC", "display_header": "FVC"}],
                    "rows": [
                        {
                            "centre_code": "SITE_A",
                            "edc_subject_ref": "SYN-001",
                            "authority_status": "not_submitted",
                            "values": {"PFT_FVC": "3.20"},
                        }
                    ],
                }
            },
            "field_mapping": [
                {
                    "event_ref": "WEEK_0",
                    "field_code": "PFT_FVC",
                    "display_header": "FVC",
                    "source_header": "FVC",
                    "revision": 0,
                }
            ],
        }
    )

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    assert workbook["导出说明"]["A1"].value == "已确认识别数据导出"
    assert [cell.value for cell in workbook["WEEK_0"][1]] == [
        "中心",
        "受试者研究编号",
        "外部EDC状态",
        "FVC",
    ]
    assert workbook["WEEK_0"]["C2"].value == "not_submitted"
    assert workbook["WEEK_0"]["D2"].value == "3.20"
