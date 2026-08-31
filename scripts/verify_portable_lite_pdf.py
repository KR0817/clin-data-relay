"""Verify the packaged Lite PDF-to-review-to-Excel workflow over localhost."""

from __future__ import annotations

import argparse
from io import BytesIO
from pathlib import Path
import secrets
import sqlite3

import httpx
from openpyxl import load_workbook

from app.offline_package import parse_encrypted_reviewed_package


EXPECTED_FIELDS = {
    "PFT_FEV1",
    "PFT_FVC",
    "PFT_FEV1_MEASURED_PREDICTED_PERCENT",
    "PFT_VT",
    "PFT_BF",
    "PFT_MV",
    "PFT_VC_MAX",
    "PFT_IC",
    "PFT_PEF",
    "PFT_MEF75",
    "PFT_MEF50",
    "PFT_MEF25",
    "PFT_VBEEX",
    "PFT_MVV",
    "PFT_TLC_SB",
    "PFT_RV_SB",
    "PFT_DLCOSB",
    "PFT_KCO",
}


def main() -> None:
    parser = argparse.ArgumentParser(description="Verify the Lite portable PDF workflow.")
    parser.add_argument("--base-url", default="http://127.0.0.1:8013")
    parser.add_argument("--pdf", type=Path, required=True)
    parser.add_argument("--image", type=Path)
    parser.add_argument("--centre-code")
    parser.add_argument("--username")
    parser.add_argument("--database", type=Path)
    args = parser.parse_args()
    if not args.pdf.is_file():
        raise RuntimeError("representative_pdf_missing")
    if args.image is not None and not args.image.is_file():
        raise RuntimeError("representative_image_missing")
    centre_mode = bool(args.centre_code or args.username or args.database)
    if centre_mode and not (args.centre_code and args.username and args.database):
        raise RuntimeError("centre_verification_arguments_required_together")

    with httpx.Client(base_url=args.base_url, timeout=90) as client:
        health = client.get("/api/health")
        health.raise_for_status()
        if health.json().get("product_mode") != "lite":
            raise RuntimeError("portable_lite_mode_not_active")

        if centre_mode:
            setup_status = client.get("/api/setup/status")
            setup_status.raise_for_status()
            expected_profile = {"centre_code": args.centre_code, "username": args.username}
            if setup_status.json() != {"required": True, "centre_profile": expected_profile}:
                raise RuntimeError("portable_centre_setup_state_mismatch")
            packaged_demo_login = client.post(
                "/api/auth/login",
                json={"username": args.username, "password": "demo-password"},
            )
            if packaged_demo_login.status_code != 401:
                raise RuntimeError("portable_centre_default_password_present")
            for forbidden_username in (
                "site-a-investigator@example.test",
                "site-b-investigator@example.test",
                "principal-investigator@example.test",
                "central-data-manager@example.test",
            ):
                if forbidden_username == args.username:
                    continue
                forbidden_login = client.post(
                    "/api/auth/login",
                    json={"username": forbidden_username, "password": "demo-password"},
                )
                if forbidden_login.status_code != 401:
                    raise RuntimeError("portable_centre_forbidden_account_present")
            password = "Aa7!" + secrets.token_urlsafe(18)
            completed = client.post(
                "/api/setup/complete",
                json={"password": password, "password_confirmation": password},
            )
            completed.raise_for_status()
            username = args.username
        else:
            username = "site-a-investigator@example.test"
            password = "demo-password"

        login = client.post(
            "/api/auth/login",
            json={"username": username, "password": password},
        )
        login.raise_for_status()
        if centre_mode and login.json()["user"]["centre_code"] != args.centre_code:
            raise RuntimeError("portable_centre_login_scope_mismatch")
        headers = {"Authorization": f"Bearer {login.json()['access_token']}"}

        if centre_mode:
            kimi_status = client.get("/api/settings/kimi", headers=headers)
            kimi_status.raise_for_status()
            if kimi_status.json().get("status") != "key_required":
                raise RuntimeError("portable_centre_kimi_initial_state_mismatch")

        if args.image is not None:
            image_upload = client.post(
                "/api/source-files/upload",
                headers=headers,
                files={"file": ("synthetic-check-sheet.png", args.image.read_bytes(), "image/png")},
                data={"synthetic_attestation": "true"},
            )
            image_upload.raise_for_status()
            draft = client.post(
                f"/api/source-files/{image_upload.json()['id']}/deidentification-drafts",
                headers=headers,
            )
            draft.raise_for_status()
            confirmed = client.post(
                f"/api/deidentification-drafts/{draft.json()['id']}/confirm",
                headers=headers,
                json={"human_review_attestation": True},
            )
            confirmed.raise_for_status()
            image_job = client.post(
                "/api/recognition-jobs",
                headers=headers,
                json={
                    "items": [
                        {
                            "source_file_id": image_upload.json()["id"],
                            "edc_subject_ref": "LITEIMG001",
                            "edc_event_ref": "WEEK_0",
                            "field_codes": ["ALT"],
                            "use_kimi": True,
                        }
                    ]
                },
            )
            image_job.raise_for_status()
            completed_image_job = client.post(
                f"/api/recognition-jobs/{image_job.json()['id']}/run",
                headers=headers,
            )
            completed_image_job.raise_for_status()
            image_candidate_ids = completed_image_job.json()["items"][0]["candidate_ids"]
            if len(image_candidate_ids) != 1:
                raise RuntimeError("portable_kimi_fallback_candidate_ids_missing")
            image_candidates = client.get("/api/candidates", headers=headers)
            image_candidates.raise_for_status()
            fallback_candidate = next(
                (
                    candidate
                    for candidate in image_candidates.json()
                    if candidate["id"] == image_candidate_ids[0]
                ),
                None,
            )
            if fallback_candidate is None or fallback_candidate["extraction_agreement"] != "local_fallback":
                raise RuntimeError("portable_kimi_fallback_provenance_mismatch")
            reviewed_image = client.post(
                "/api/candidate-reviews/bulk-accept",
                headers=headers,
                json={"candidate_ids": image_candidate_ids},
            )
            reviewed_image.raise_for_status()
            if reviewed_image.json()["accepted_count"] != 1:
                raise RuntimeError("portable_kimi_fallback_bulk_review_failed")

        if centre_mode:
            placeholder_key = "sk-package-qa-" + secrets.token_urlsafe(24)
            configured_kimi = client.put(
                "/api/settings/kimi",
                headers=headers,
                json={"key": placeholder_key},
            )
            configured_kimi.raise_for_status()
            if configured_kimi.json() != {
                "configured": True,
                "status": "ready",
                "provider": "kimi",
                "model": "kimi-k3",
            } or placeholder_key in configured_kimi.text:
                raise RuntimeError("portable_centre_kimi_web_configuration_failed")

        options = client.get("/api/recognition-fields?event_ref=WEEK_0", headers=headers)
        options.raise_for_status()
        pulmonary_codes = {
            item["field_code"]
            for item in options.json()["fields"]
            if item["category"] == "pulmonary_function"
        }
        if pulmonary_codes != EXPECTED_FIELDS:
            raise RuntimeError("portable_lite_pulmonary_dictionary_mismatch")

        upload = client.post(
            "/api/source-files/upload",
            headers=headers,
            files={"file": ("representative-pulmonary-report.pdf", args.pdf.read_bytes(), "application/pdf")},
            data={
                "synthetic_attestation": "true",
                "edc_subject_ref": "LITEPFT001",
                "edc_event_ref": "WEEK_0",
            },
        )
        upload.raise_for_status()
        recognition_job = client.post(
            "/api/recognition-jobs",
            headers=headers,
            json={
                "items": [
                    {
                        "source_file_id": upload.json()["id"],
                        "edc_subject_ref": "LITEPFT001",
                        "edc_event_ref": "WEEK_0",
                        "field_codes": sorted(EXPECTED_FIELDS),
                        "use_kimi": False,
                    }
                ]
            },
        )
        recognition_job.raise_for_status()
        completed_job = client.post(
            f"/api/recognition-jobs/{recognition_job.json()['id']}/run",
            headers=headers,
        )
        completed_job.raise_for_status()
        candidate_ids = completed_job.json()["items"][0]["candidate_ids"]
        restored_job = client.get(
            f"/api/recognition-jobs/{recognition_job.json()['id']}",
            headers=headers,
        )
        restored_job.raise_for_status()
        if restored_job.json()["items"][0]["candidate_ids"] != candidate_ids:
            raise RuntimeError("portable_lite_recognition_batch_not_restored")
        listed_candidates = client.get("/api/candidates", headers=headers)
        listed_candidates.raise_for_status()
        candidates = [
            candidate
            for candidate in listed_candidates.json()
            if candidate["id"] in set(candidate_ids)
        ]
        actual_fields = {candidate["field_code"] for candidate in candidates}
        if actual_fields != EXPECTED_FIELDS:
            raise RuntimeError("portable_lite_pdf_field_mismatch")

        reviewed = client.post(
            "/api/candidate-reviews/bulk-accept",
            headers=headers,
            json={"candidate_ids": candidate_ids},
        )
        reviewed.raise_for_status()
        if reviewed.json()["accepted_count"] != len(EXPECTED_FIELDS):
            raise RuntimeError("portable_lite_bulk_review_failed")

        workbook_response = client.get(
            "/api/exports/reviewed-recognition-data.xlsx",
            headers=headers,
        )
        workbook_response.raise_for_status()
        workbook = load_workbook(BytesIO(workbook_response.content), read_only=True, data_only=True)
        if "WEEK_0" not in workbook.sheetnames:
            raise RuntimeError("portable_lite_excel_event_sheet_missing")
        rows = list(workbook["WEEK_0"].iter_rows(values_only=True))
        expected_data_rows = 1 + int(args.image is not None)
        if len(rows) != expected_data_rows + 1:
            raise RuntimeError("portable_lite_excel_row_count_mismatch")
        headers_row = {str(value) for value in rows[0] if value is not None}
        if not {"FEV1", "FVC", "DLCOSB", "KCO"} <= headers_row:
            raise RuntimeError("portable_lite_excel_fields_missing")
        if "not_submitted" not in {str(value) for value in rows[1] if value is not None}:
            raise RuntimeError("portable_lite_excel_authority_state_missing")

        if centre_mode:
            package_passphrase = "Package-" + secrets.token_urlsafe(18)
            package_response = client.post(
                "/api/exports/reviewed-recognition-package.json",
                headers=headers,
                data={"package_passphrase": package_passphrase},
            )
            package_response.raise_for_status()
            package, package_sha256 = parse_encrypted_reviewed_package(
                package_response.content,
                passphrase=package_passphrase,
            )
            expected_record_count = len(EXPECTED_FIELDS) + int(args.image is not None)
            if package["centre_code"] != args.centre_code or package["record_count"] != expected_record_count:
                raise RuntimeError("portable_centre_package_scope_mismatch")
            if package_response.headers.get("x-offline-package-sha256") != package_sha256:
                raise RuntimeError("portable_centre_package_hash_mismatch")

    if centre_mode:
        connection = sqlite3.connect(f"file:{args.database.resolve()}?mode=ro", uri=True)
        users = connection.execute(
            "SELECT username, centre_code, role, active, password_hash FROM users"
        ).fetchall()
        connection.close()
        if len(users) != 1 or users[0][:4] != (args.username, args.centre_code, "site_investigator", 1):
            raise RuntimeError("portable_centre_repository_scope_mismatch")
        if not str(users[0][4]).startswith("scrypt$"):
            raise RuntimeError("portable_centre_password_hash_mismatch")

    if centre_mode:
        print("PASS: Centre package verified Kimi fallback, reviewed 19 fields and exported an encrypted centre package.")
    elif args.image is not None:
        print("PASS: Lite package verified local fallback, reviewed 19 fields and exported Excel.")
    else:
        print("PASS: Lite package parsed 18 pulmonary fields, recorded human review and exported Excel.")


if __name__ == "__main__":
    main()
