"""Isolated Artifact Tool boundary for submitted-data Excel exports."""

from __future__ import annotations

import json
import os
import subprocess
import tempfile
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class SpreadsheetExportError(RuntimeError):
    """Raised when the external workbook builder cannot produce an export."""


class ArtifactToolSpreadsheetExporter:
    def __init__(self, *, node_executable: Path | None, builder_script: Path, output_directory: Path) -> None:
        self.node_executable = node_executable
        self.builder_script = builder_script
        self.output_directory = output_directory

    @classmethod
    def from_environment(cls) -> "ArtifactToolSpreadsheetExporter":
        project_root = Path(__file__).resolve().parent.parent
        configured_node = os.getenv("SPREADSHEET_NODE_EXECUTABLE", "").strip()
        return cls(
            node_executable=Path(configured_node) if configured_node else None,
            builder_script=project_root / "scripts" / "build_submitted_export.mjs",
            output_directory=project_root / ".runtime" / "exports",
        )

    @property
    def artifact_tool_ready(self) -> bool:
        return bool(
            self.node_executable is not None
            and self.node_executable.is_file()
            and self.builder_script.is_file()
        )

    @property
    def ready(self) -> bool:
        return True

    def export(self, payload: dict[str, object]) -> bytes:
        if not self.artifact_tool_ready or self.node_executable is None:
            return self._export_with_openpyxl(payload)
        self.output_directory.mkdir(parents=True, exist_ok=True)
        handle = tempfile.NamedTemporaryFile(
            prefix="submitted-data-",
            suffix=".xlsx",
            dir=self.output_directory,
            delete=False,
        )
        output_path = Path(handle.name)
        handle.close()
        try:
            completed = subprocess.run(  # nosec B603: executable and script are resolved server configuration
                [str(self.node_executable), str(self.builder_script), str(output_path)],
                input=json.dumps(payload, ensure_ascii=False),
                text=True,
                encoding="utf-8",
                capture_output=True,
                timeout=60,
                check=False,
            )
            if completed.returncode != 0 or not output_path.is_file():
                raise SpreadsheetExportError("spreadsheet_export_failed")
            workbook_bytes = output_path.read_bytes()
            if not workbook_bytes.startswith(b"PK"):
                raise SpreadsheetExportError("spreadsheet_export_invalid_file")
            return workbook_bytes
        except (OSError, subprocess.SubprocessError) as error:
            raise SpreadsheetExportError("spreadsheet_export_failed") from error
        finally:
            output_path.unlink(missing_ok=True)

    @staticmethod
    def _excel_safe_value(value: object) -> object:
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value

    @staticmethod
    def _mapping(value: object) -> dict[str, object]:
        if not isinstance(value, dict):
            raise SpreadsheetExportError("spreadsheet_export_invalid_payload")
        return value

    @staticmethod
    def _list(value: object) -> list[object]:
        if not isinstance(value, list):
            raise SpreadsheetExportError("spreadsheet_export_invalid_payload")
        return value

    @classmethod
    def _export_with_openpyxl(cls, payload: dict[str, object]) -> bytes:
        dark_blue = "1E40AF"
        pale_blue = "DBEAFE"
        ink = "172554"
        white = "FFFFFF"

        try:
            export_title = str(payload.get("export_title") or "已提交临床数据导出")
            inclusion_rule = str(
                payload.get("inclusion_rule") or "仅纳入 LibreClinica 已确认 submitted 的记录"
            )
            authority_note = str(
                payload.get("authority_note") or "LibreClinica；本工作簿为便捷导出副本"
            )
            value_count = payload.get("value_count", payload.get("submitted_value_count", 0))
            include_authority_status = payload.get("include_authority_status") is True
            authority_status_header = str(payload.get("authority_status_header") or "LibreClinica状态")
            workbook = Workbook()
            notes = workbook.active
            notes.title = "导出说明"
            notes.sheet_view.showGridLines = False
            notes_rows = [
                [export_title, ""],
                ["生成时间（UTC）", f"UTC {payload.get('generated_at', '')}"],
                ["导出范围", payload.get("scope", "")],
                ["纳入规则", inclusion_rule],
                ["权威记录", authority_note],
                ["数据字典", payload.get("dictionary_id", "")],
                ["字典版本", payload.get("dictionary_version", "")],
                ["字段值数量", value_count],
            ]
            for row in notes_rows:
                notes.append(row)
            notes.merge_cells("A1:B1")
            notes["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color=white)
            notes["A1"].fill = PatternFill("solid", fgColor=dark_blue)
            notes["A1"].alignment = Alignment(vertical="center")
            for row_index in range(2, 9):
                notes.cell(row=row_index, column=1).font = Font(name="Microsoft YaHei", size=10, bold=True, color=ink)
                notes.cell(row=row_index, column=1).fill = PatternFill("solid", fgColor=pale_blue)
                notes.cell(row=row_index, column=2).font = Font(name="Microsoft YaHei", size=10, color=ink)
                notes.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True, vertical="center")
            notes.column_dimensions["A"].width = 24
            notes.column_dimensions["B"].width = 58
            notes.row_dimensions[1].height = 30

            events = cls._mapping(payload.get("events", {}))
            for event_ref, raw_event_data in events.items():
                event_data = cls._mapping(raw_event_data)
                columns = [cls._mapping(column) for column in cls._list(event_data.get("columns", []))]
                rows = [cls._mapping(row) for row in cls._list(event_data.get("rows", []))]
                sheet = workbook.create_sheet(title=str(event_ref)[:31])
                sheet.sheet_view.showGridLines = False
                fixed_headers = ["中心", "受试者研究编号"]
                if include_authority_status:
                    fixed_headers.append(cls._excel_safe_value(authority_status_header))
                headers = fixed_headers + [
                    cls._excel_safe_value(column.get("display_header", "")) for column in columns
                ]
                sheet.append(headers)
                for row in rows:
                    values = cls._mapping(row.get("values", {}))
                    fixed_values = [
                        cls._excel_safe_value(row.get("centre_code", "")),
                        cls._excel_safe_value(row.get("edc_subject_ref", "")),
                    ]
                    if include_authority_status:
                        fixed_values.append(cls._excel_safe_value(row.get("authority_status", "")))
                    sheet.append(fixed_values + [
                        cls._excel_safe_value(values.get(str(column.get("field_code", ""))))
                        for column in columns
                    ])
                cls._style_data_sheet(
                    sheet,
                    dark_blue=dark_blue,
                    white=white,
                    ink=ink,
                    fixed_columns=len(fixed_headers),
                )

            mapping_sheet = workbook.create_sheet(title="字段映射")
            mapping_sheet.sheet_view.showGridLines = False
            mapping_sheet.append(["访视", "字段代码（不可变）", "当前显示表头", "原始 Excel 表头", "修订版本"])
            for raw_item in cls._list(payload.get("field_mapping", [])):
                item = cls._mapping(raw_item)
                mapping_sheet.append(
                    [
                        cls._excel_safe_value(item.get("event_ref", "")),
                        cls._excel_safe_value(item.get("field_code", "")),
                        cls._excel_safe_value(item.get("display_header", "")),
                        cls._excel_safe_value(item.get("source_header", "")),
                        item.get("revision", ""),
                    ]
                )
            cls._style_data_sheet(mapping_sheet, dark_blue=dark_blue, white=white, ink=ink)
            for column_index, width in enumerate((22, 22, 34, 34, 14), start=1):
                mapping_sheet.column_dimensions[get_column_letter(column_index)].width = width

            output = BytesIO()
            workbook.save(output)
            workbook_bytes = output.getvalue()
            if not workbook_bytes.startswith(b"PK"):
                raise SpreadsheetExportError("spreadsheet_export_invalid_file")
            return workbook_bytes
        except SpreadsheetExportError:
            raise
        except (KeyError, OSError, TypeError, ValueError) as error:
            raise SpreadsheetExportError("spreadsheet_export_failed") from error

    @staticmethod
    def _style_data_sheet(
        sheet: object,
        *,
        dark_blue: str,
        white: str,
        ink: str,
        fixed_columns: int = 2,
    ) -> None:
        sheet.freeze_panes = f"{get_column_letter(fixed_columns + 1)}2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 36
        for cell in sheet[1]:
            cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=dark_blue)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Microsoft YaHei", size=10, color=ink)
                cell.alignment = Alignment(vertical="center")
        for column_index in range(1, sheet.max_column + 1):
            width = 16 if column_index == 1 else 22 if column_index == 2 else 20 if column_index <= fixed_columns else 18
            sheet.column_dimensions[get_column_letter(column_index)].width = width
